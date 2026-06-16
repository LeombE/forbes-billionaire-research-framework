"""Excel workbook export."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .config import DEFAULT_TARGET_YEAR, YearConfig, get_year_config


SHEET_TABLE_NAMES = {
    "Top100_2025": "Top100Table",
    "Wealth_History_Long": "WealthHistoryTable",
    "Growth_Metrics": "GrowthMetricsTable",
    "Industry_Summary": "IndustrySummaryTable",
    "Country_Summary": "CountrySummaryTable",
    "Wealth_Engine_Summary": "WealthEngineSummaryTable",
    "Source_Citations": "SourceCitationsTable",
    "Person_Data_Quality": "PersonDataQualityTable",
    "Data_Quality": "DataQualityTable",
}


def _sheet_table_names(config: YearConfig) -> dict[str, str]:
    if config.legacy_layout:
        return SHEET_TABLE_NAMES
    return {
        config.top100_sheet: f"Top100{config.year}Table",
        config.wealth_history_sheet: f"WealthHistory{config.year}Table",
        config.growth_metrics_sheet: f"GrowthMetrics{config.year}Table",
        config.industry_summary_sheet: f"IndustrySummary{config.year}Table",
        config.country_summary_sheet: f"CountrySummary{config.year}Table",
        config.wealth_engine_summary_sheet: f"WealthEngineSummary{config.year}Table",
        config.source_citations_sheet: f"SourceCitations{config.year}Table",
        config.evidence_registry_sheet: f"EvidenceRegistry{config.year}Table",
        config.data_quality_sheet: f"DataQuality{config.year}Table",
    }


def _methodology_frame(methodology_text: str) -> pd.DataFrame:
    lines = [line if line.strip() else " " for line in methodology_text.splitlines()]
    return pd.DataFrame({"Methodology": lines})


def _add_table(ws, table_name: str) -> None:
    if ws.max_row < 2 or ws.max_column < 1:
        return
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)


def _set_column_widths(ws) -> None:
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        header = str(column_cells[0].value or "")
        max_length = len(header)
        for cell in list(column_cells)[1:301]:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value), 80))
        width = max(10, min(max_length + 2, 55))
        if header in {"evidence_summary", "notes", "source_url", "forbes_profile_url", "canonical_source_url"}:
            width = 50
        ws.column_dimensions[column_letter].width = width


def _format_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            header = str(ws.cell(row=1, column=cell.column).value or "")
            if header.endswith("_usd_b") or "net_worth" in header or "gain_usd_b" in header or "loss_usd_b" in header:
                cell.number_format = '#,##0.0'
            elif "CAGR" in header or header.endswith("_pct") or "growth" in header:
                cell.number_format = "0.0%"
            elif "r2" in header.lower() or "slope" in header.lower() or "score" in header.lower():
                cell.number_format = "0.000"
            elif "year" in header.lower() or "rank" in header.lower() or "age" in header.lower() or "count" in header.lower():
                cell.number_format = "0"
            if header in {"evidence_summary", "notes", "Methodology"}:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")
    _set_column_widths(ws)


def _add_summary_chart(ws, category_col: int, value_col: int, title: str, anchor: str = "H2") -> None:
    if ws.max_row < 3:
        return
    end_row = min(ws.max_row, 13)
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = ""
    chart.x_axis.title = str(ws.cell(row=1, column=value_col).value or "")
    data = Reference(ws, min_col=value_col, min_row=1, max_row=end_row)
    cats = Reference(ws, min_col=category_col, min_row=2, max_row=end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 12
    ws.add_chart(chart, anchor)


def write_excel_workbook(
    output_path: Path,
    top100: pd.DataFrame,
    history_long: pd.DataFrame,
    metrics: pd.DataFrame,
    summaries: dict[str, pd.DataFrame],
    source_citations: pd.DataFrame,
    person_quality: pd.DataFrame,
    quality_checks: pd.DataFrame,
    methodology_text: str,
    config: YearConfig | None = None,
    evidence_registry: pd.DataFrame | None = None,
) -> None:
    """Write the required multi-sheet Excel workbook."""
    config = config or get_year_config(DEFAULT_TARGET_YEAR)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    sheets = {
        config.top100_sheet: top100,
        config.wealth_history_sheet: history_long,
        config.growth_metrics_sheet: metrics,
        config.industry_summary_sheet: summaries["Industry_Summary"],
        config.country_summary_sheet: summaries["Country_Summary"],
        config.wealth_engine_summary_sheet: summaries["Wealth_Engine_Summary"],
        config.source_citations_sheet: source_citations,
        config.data_quality_sheet: quality_checks,
        config.methodology_sheet: _methodology_frame(methodology_text),
    }
    if config.legacy_layout:
        sheets["Person_Data_Quality"] = person_quality
    else:
        sheets[config.evidence_registry_sheet] = evidence_registry if evidence_registry is not None else pd.DataFrame()

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    workbook = load_workbook(output_path)
    table_names = _sheet_table_names(config)
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        _format_sheet(ws)
        if sheet_name in table_names:
            _add_table(ws, table_names[sheet_name])

    _add_summary_chart(
        workbook[config.industry_summary_sheet],
        category_col=1,
        value_col=3,
        title=f"Top Industries by Total {config.year} Net Worth",
        anchor="H2",
    )
    _add_summary_chart(
        workbook[config.country_summary_sheet],
        category_col=1,
        value_col=2,
        title="Top Countries/Territories by People Count",
        anchor="H2",
    )
    _add_summary_chart(
        workbook[config.wealth_engine_summary_sheet],
        category_col=1,
        value_col=2,
        title="Wealth Engine Count",
        anchor="I2",
    )

    workbook.save(output_path)
