"""Command-line pipeline for the multi-year Forbes Top 100 analysis."""

from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from .citations import build_person_data_quality_scores, build_source_citations
from .clean import build_top100, build_top100_2025, build_wealth_history_long
from .config import (
    BASE_DIR,
    DEFAULT_TARGET_YEAR,
    YearConfig,
    annual_api_url,
    ensure_year_dirs,
    get_year_config,
)
from .excel import write_excel_workbook
from .fetch_forbes import ForbesAnnualFetcher
from .manual_templates import load_manual_citations, load_manual_inputs, write_manual_import_templates
from .metrics import calculate_growth_metrics
from .quality import missing_field_summary, run_quality_checks, write_quality_report
from .reports import create_charts, write_executive_summary, write_insight_report
from .summaries import build_all_summaries


def _load_cached_payloads(years: list[int], config: YearConfig) -> dict[int, dict]:
    fetcher = ForbesAnnualFetcher(config=config)
    payloads: dict[int, dict] = {}
    for year in years:
        payloads[year] = fetcher.load_cached_year(year)
    return payloads


def _build_methodology_text(
    successful_years: list[int],
    fetch_failures: list[str],
    robots_note: str,
    manual_mode: bool,
    config: YearConfig,
) -> str:
    missing_years = [year for year in config.history_years if year not in successful_years]
    source_mode = "manual import files" if manual_mode else "official Forbes annual-list JSON endpoints"
    raw_error_logs = sorted(path.name for path in config.raw_forbes_dir.glob("*_error.txt"))
    raw_error_note = (
        "none"
        if not raw_error_logs
        else (
            f"{len(raw_error_logs)} preserved diagnostic file(s): {', '.join(raw_error_logs)}. "
            "These are fetch diagnostics and do not override successful cached JSON snapshots."
        )
    )
    return f"""# Methodology Notes

## Source Scope

- Canonical population: Forbes World's Billionaires {config.year} annual list, top 100 rows by Forbes position.
- Historical source mode: {source_mode}.
- Forbes Real-Time Billionaires is intentionally excluded from the canonical annual dataset.
- Forbes profile URLs are stored for audit, but profile prose is not copied into processed reports.
- Forbes annual API URL pattern: `{annual_api_url(config.year)}`.
- Robots/permissions note: {robots_note}

## Annual History Coverage

- Successful annual snapshots: {', '.join(map(str, successful_years)) if successful_years else 'none'}.
- Missing or failed annual snapshots: {', '.join(map(str, missing_years)) if missing_years else 'none'}.
- Fetch failures: {'; '.join(fetch_failures) if fetch_failures else 'none'}.
- Preserved raw fetch diagnostic logs: {raw_error_note}

## Cleaning Rules

- `finalWorth` is converted from Forbes API units of USD millions to USD billions.
- People are matched across years by Forbes profile URI.
- Forbes rows labelled "& family" are retained as Forbes publishes them; this project does not split family fortunes.
- The rank field preserves Forbes rank, including ties. The position field stores the ordered top-100 position and is unique.
- Missing values are left blank/NA. The project does not fabricate unknown ages, countries, sources, or histories.

## Growth Formulas

- Wealth multiple: target-year net worth divided by first observed net worth.
- Nominal CAGR: `(target_year_net_worth / first_net_worth) ** (1 / years_elapsed) - 1`.
- Exponential-style fit: `ln(net_worth_usd_b) = a + b * year`.
- Doubling time: `ln(2) / b` only when `b > 0`.
- CAGR, wealth multiple, one-year gain/loss, volatility, drawdown, log-linear slope, R^2, and doubling time require at least three positive annual observations.
- Do not claim true exponential growth unless data coverage and fit quality support it.

## Source Policy

- Forbes annual-list data is canonical for annual rank, annual net worth, country, industry, source of wealth, and profile URL.
- Forbes Real-Time Billionaires data may only appear in a separate comparison file and must not overwrite annual fields.
- If official annual data is blocked, paywalled, rate-limited, or otherwise unavailable, the project uses manual-import templates and leaves missing values blank.

## Wealth Engine Classification

The classifier is transparent and rule-based. Each person receives exactly one dominant category from the allowed list. Categories can overlap in reality, so the assigned label should be read as the primary rule match, not as a complete biography.

## Limitations

- Forbes net worth values are estimates and annual snapshots, not audited personal balance sheets.
- Historical URI matching can miss a person if Forbes changed profile identifiers.
- Public-market re-ratings, IPOs, donations, divorces, inheritance changes, and FX movements can create jumps that are not smooth compounding.
- The workbook and reports are for research and strategy analysis; review Forbes terms before redistributing raw source data.
"""


def _empty_evidence_registry(config: YearConfig) -> pd.DataFrame:
    columns = [
        "report_year",
        "forbes_uri",
        "person_slug",
        "person_name",
        "rank",
        "route_id",
        "template_year",
        "template_reference_only",
        "source_key",
        "claim_year",
        "source_as_of_date",
        "claim_supported",
        "source_title",
        "source_url_or_locator",
        "confidence",
        "limitation",
    ]
    return pd.DataFrame(columns=columns)


def _write_outputs(
    top100: pd.DataFrame,
    history_long: pd.DataFrame,
    metrics: pd.DataFrame,
    source_citations: pd.DataFrame,
    person_quality: pd.DataFrame,
    summaries: dict[str, pd.DataFrame],
    quality_checks: pd.DataFrame,
    methodology_text: str,
    config: YearConfig,
    evidence_registry: pd.DataFrame,
) -> None:
    config.processed_dir.mkdir(parents=True, exist_ok=True)
    config.annual_reports_dir.mkdir(parents=True, exist_ok=True)
    top100.to_csv(config.top100_path, index=False)
    history_long.to_csv(config.history_path, index=False)
    metrics.to_csv(config.metrics_path, index=False)
    source_citations.to_csv(config.citations_path, index=False)
    person_quality.to_csv(config.person_quality_path, index=False)

    for name, df in summaries.items():
        suffix = "" if config.legacy_layout else f"_{config.year}"
        df.to_csv(config.processed_dir / f"{name.lower()}{suffix}.csv", index=False)

    methodology_path = BASE_DIR / "methodology_notes.md" if config.legacy_layout else config.annual_reports_dir / f"methodology_notes_{config.year}.md"
    methodology_path.write_text(methodology_text, encoding="utf-8")

    if not config.legacy_layout:
        evidence_registry.to_csv(config.interim_dir / f"enriched_evidence_registry_{config.year}.csv", index=False)

    write_executive_summary(top100, metrics, summaries, config=config)
    write_insight_report(top100, history_long, metrics, summaries, config=config)
    create_charts(top100, history_long, metrics, summaries, config=config)
    write_excel_workbook(
        config.workbook_path,
        top100,
        history_long,
        metrics,
        summaries,
        source_citations,
        person_quality,
        quality_checks,
        methodology_text,
        config=config,
        evidence_registry=evidence_registry,
    )


def _verify_workbook(path: Path, config: YearConfig) -> None:
    expected_sheets = (
        {
            "Top100_2025",
            "Wealth_History_Long",
            "Growth_Metrics",
            "Industry_Summary",
            "Country_Summary",
            "Wealth_Engine_Summary",
            "Source_Citations",
            "Person_Data_Quality",
            "Data_Quality",
            "Methodology",
        }
        if config.legacy_layout
        else {
            config.top100_sheet,
            config.wealth_history_sheet,
            config.growth_metrics_sheet,
            config.industry_summary_sheet,
            config.country_summary_sheet,
            config.wealth_engine_summary_sheet,
            config.source_citations_sheet,
            config.evidence_registry_sheet,
            config.data_quality_sheet,
            config.methodology_sheet,
        }
    )
    workbook = load_workbook(path, read_only=True)
    missing = expected_sheets - set(workbook.sheetnames)
    if missing:
        raise RuntimeError(f"Workbook is missing required sheets: {sorted(missing)}")


def _load_or_fetch_payloads(config: YearConfig, *, offline: bool, force_fetch: bool) -> tuple[dict[int, dict], list[str], str]:
    if offline:
        return _load_cached_payloads(config.history_years, config), [], "Offline mode; robots.txt was not fetched during this run."

    fetcher = ForbesAnnualFetcher(config=config)
    payloads, fetch_failures, robots_check = fetcher.fetch_years(config.history_years, force=force_fetch)
    return payloads, fetch_failures, robots_check.note


def run_pipeline(
    *,
    year: int = DEFAULT_TARGET_YEAR,
    force_fetch: bool = False,
    offline: bool = False,
    manual: bool = False,
    manual_import_dir: str | Path | None = None,
    init_year: bool = False,
) -> dict[str, object]:
    config = get_year_config(year)
    ensure_year_dirs(config)
    template_paths = write_manual_import_templates(config.year)

    if init_year:
        return {
            "year": config.year,
            "mode": "init_year",
            "manual_import_required": not config.legacy_layout,
            "templates": {key: str(value) for key, value in template_paths.items() if key.endswith("template")},
        }

    fetch_failures: list[str] = []
    robots_note = "Manual/offline mode; robots.txt was not fetched during this run."

    if manual:
        top100, history_long = load_manual_inputs(config.year, manual_import_dir=manual_import_dir)
        if top100.empty or history_long.empty:
            manual_source = Path(manual_import_dir) if manual_import_dir is not None else Path("templates")
            raise RuntimeError(
                f"Manual import files for {config.year} are present but empty in {manual_source}. Fill official "
                "Forbes annual-list manual-import files before running the pipeline in manual mode."
            )
        if "_classification_text" not in top100.columns:
            top100["_classification_text"] = ""
        successful_years = sorted(history_long["year"].dropna().astype(int).unique().tolist())
        top100_for_metrics = top100.copy()
    else:
        payloads, fetch_failures, robots_note = _load_or_fetch_payloads(config, offline=offline, force_fetch=force_fetch)
        if config.year not in payloads:
            raise RuntimeError(
                f"Missing canonical Forbes {config.year} annual list. If Forbes access is blocked, fill the "
                "year-specific manual-import templates, then run with --manual-import."
            )
        successful_years = sorted(payloads.keys())
        top100_with_internal = (
            build_top100_2025(payloads[config.year]) if config.legacy_layout else build_top100(payloads[config.year], config)
        )
        history_long = build_wealth_history_long(payloads, top100_with_internal, config)
        top100 = top100_with_internal.drop(columns=["_classification_text"])
        top100_for_metrics = top100_with_internal

    metrics = calculate_growth_metrics(top100_for_metrics, history_long, successful_years, config)
    if manual:
        manual_citations = load_manual_citations(config.year, manual_import_dir=manual_import_dir)
        if manual_citations is None or manual_citations.empty:
            manual_source = Path(manual_import_dir) if manual_import_dir is not None else Path("templates")
            raise RuntimeError(
                f"Manual mode for {config.year} requires a filled manual source citation file in {manual_source}; "
                "blank templates are not enough."
            )
        source_citations = manual_citations
    else:
        source_citations = build_source_citations(top100, history_long, metrics, successful_years, config)

    person_quality = build_person_data_quality_scores(top100, history_long, metrics, source_citations, config)
    summaries = build_all_summaries(top100, metrics, config)
    quality_checks = run_quality_checks(top100, history_long, metrics, source_citations, fetch_failures, person_quality, config)
    missing_summary = missing_field_summary(top100, metrics)
    methodology_text = _build_methodology_text(successful_years, fetch_failures, robots_note, manual, config)
    evidence_registry = _empty_evidence_registry(config)

    pd.DataFrame({"successful_year": successful_years}).to_csv(config.raw_forbes_dir / "successful_years.csv", index=False)
    quality_checks.to_csv(config.processed_dir / config.data_quality_checks_filename, index=False)
    missing_summary.to_csv(config.processed_dir / config.missing_summary_filename, index=False)

    quality_report_path = BASE_DIR / "data_quality_report.md" if config.legacy_layout else config.annual_reports_dir / f"data_quality_report_{config.year}.md"
    write_quality_report(quality_report_path, quality_checks, missing_summary, top100, metrics, person_quality, config)
    _write_outputs(
        top100,
        history_long,
        metrics,
        source_citations,
        person_quality,
        summaries,
        quality_checks,
        methodology_text,
        config,
        evidence_registry,
    )
    _verify_workbook(config.workbook_path, config)

    failures = quality_checks[quality_checks["status"].eq("FAIL")]
    if not failures.empty:
        raise RuntimeError("Quality checks failed:\n" + failures.to_string(index=False))

    return {
        "year": config.year,
        "top100_rows": len(top100),
        "history_rows": len(history_long),
        "metrics_rows": len(metrics),
        "citation_rows": len(source_citations),
        "person_quality_rows": len(person_quality),
        "successful_years": successful_years,
        "fetch_failures": fetch_failures,
        "workbook": str(config.workbook_path),
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run the multi-year Forbes Top 100 billionaire research pipeline.")
    parser.add_argument("--year", type=int, default=DEFAULT_TARGET_YEAR, help="Canonical annual-list year to process.")
    parser.add_argument("--force-fetch", action="store_true", help="Refetch Forbes annual JSON even if raw cache exists.")
    parser.add_argument("--offline", action="store_true", help="Use cached Forbes annual JSON files; do not fetch.")
    parser.add_argument("--manual", "--manual-import", action="store_true", help="Use year-specific manual CSV inputs.")
    parser.add_argument(
        "--manual-import-dir",
        type=Path,
        default=None,
        help="Read manual-import CSVs from a private local directory instead of public templates/.",
    )
    parser.add_argument("--init-year", action="store_true", help="Create year directories and manual-import templates only.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    result = run_pipeline(
        year=args.year,
        force_fetch=args.force_fetch,
        offline=args.offline,
        manual=args.manual,
        manual_import_dir=args.manual_import_dir,
        init_year=args.init_year,
    )
    print("Pipeline complete")
    for key, value in result.items():
        print(f"{key}: {value}")


if __name__ == "__main__":
    main()
