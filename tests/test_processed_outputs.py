from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.config import CITATION_COLUMNS, WEALTH_ENGINE_CATEGORIES
from src.citations import TOP100_CITED_FIELDS
from src.docx_reports import build_report_context, create_person_report, validate_docx_report


BASE_DIR = Path(__file__).resolve().parents[1]


def test_required_processed_outputs_after_pipeline() -> None:
    top100_path = BASE_DIR / "data" / "processed" / "top100_2025.csv"
    if not top100_path.exists():
        pytest.skip("Pipeline outputs not generated yet.")

    top100 = pd.read_csv(top100_path)
    metrics = pd.read_csv(BASE_DIR / "data" / "processed" / "billionaire_growth_metrics.csv")
    citations = pd.read_csv(BASE_DIR / "data" / "processed" / "source_citations.csv")
    person_quality = pd.read_csv(BASE_DIR / "data" / "processed" / "person_data_quality_scores.csv")

    assert len(top100) == 100
    assert top100["forbes_uri"].nunique() == 100
    assert top100["net_worth_2025_usd_b"].notna().all()
    assert top100["canonical_source_url"].astype(str).str.len().gt(0).all()
    assert len(metrics) == 100
    assert set(CITATION_COLUMNS).issubset(citations.columns)
    assert citations["source_id"].astype(str).str.len().gt(0).all()
    assert citations["source_id"].is_unique
    non_derived = citations[citations["collection_method"].ne("derived_metric")]
    assert non_derived["source_url"].astype(str).str.len().gt(0).all()
    assert metrics["wealth_engine_category"].isin(WEALTH_ENGINE_CATEGORIES).all()
    assert metrics.loc[metrics["years_observed"] < 3, "CAGR_nominal"].isna().all()
    assert metrics.loc[metrics["years_observed"] < 3, "exponential_fit_r2"].isna().all()
    assert len(person_quality) == 100


def test_source_citation_field_coverage_after_pipeline() -> None:
    top100_path = BASE_DIR / "data" / "processed" / "top100_2025.csv"
    if not top100_path.exists():
        pytest.skip("Pipeline outputs not generated yet.")

    top100 = pd.read_csv(top100_path)
    citations = pd.read_csv(BASE_DIR / "data" / "processed" / "source_citations.csv")
    keys = set(citations[["table_name", "forbes_uri", "field_name"]].fillna("").astype(str).agg("|".join, axis=1))
    missing = []
    for _, row in top100.iterrows():
        for field in TOP100_CITED_FIELDS:
            key = f"top100_2025|{row['forbes_uri']}|{field}"
            if key not in keys:
                missing.append(key)
    assert not missing


def test_workbook_required_sheets_after_pipeline() -> None:
    workbook_path = BASE_DIR / "Forbes_top100_2025_analysis.xlsx"
    if not workbook_path.exists():
        pytest.skip("Workbook not generated yet.")

    workbook = load_workbook(workbook_path, read_only=True)
    expected = {
        "Top100_2025",
        "Wealth_History_Long",
        "Growth_Metrics",
        "Industry_Summary",
        "Country_Summary",
        "Wealth_Engine_Summary",
        "Source_Citations",
        "Person_Data_Quality",
        "Data_Quality",
        "Methodology",
    }
    assert expected.issubset(set(workbook.sheetnames))


def test_docx_report_generator_creates_cited_quality_checked_report(tmp_path: Path) -> None:
    if not (BASE_DIR / "data" / "processed" / "top100_2025.csv").exists():
        pytest.skip("Pipeline outputs not generated yet.")

    context = build_report_context(rank=1)
    output_path = create_person_report(context, output_dir=tmp_path)
    validation = validate_docx_report(output_path)

    assert output_path.exists()
    assert validation["passed"], validation
